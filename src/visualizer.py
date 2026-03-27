from __future__ import annotations

import re
from pathlib import Path
from typing import List, Optional, Tuple

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import numpy as np
import pandas as pd
from src.utils import format_number_kmg, format_integer_commas
import seaborn as sns
from matplotlib import font_manager, rcParams
from rich.console import Console
from rich.table import Table

# --- Add/Replace inside _smart_label_anchor ---
# --- REPLACED: smarter label anchor with explicit side & vertical preference ---
import matplotlib.dates as _mdates

def _smart_label_anchor(
    ax,
    x_data,
    y_data,
    *,
    final_xy=None,
    dx=8,
    dy=10,
    force_side: str = None,  # 'left' | 'right' | None
    prefer_v: str = None,    # 'up'   | 'down'  | None
):
    """
    Decide (xytext, ha, va) so that the text stays inside axes and avoids
    collision with the final value tag when it is too close.

    New behavior:
      - force_side: force the horizontal side of the label ("left" or "right")
      - prefer_v  : bias vertical placement ("up" for drawup, "down" for drawdown)
    This function auto-flips near edges to keep labels inside the axes.
    """

    # Axis limits in data units
    x0, x1 = ax.get_xlim()
    y0, y1 = ax.get_ylim()

    # Convert x (and final x) to numeric axis units (e.g., date -> float days)
    def _x_to_num(x):
        try:
            return _mdates.date2num(x)
        except Exception:
            try:
                return float(x)
            except Exception:
                return x0 + 0.5 * (x1 - x0)

    xd = _x_to_num(x_data)
    fx_num = _x_to_num(final_xy[0]) if final_xy is not None else None

    xr = 0.5 if x1 == x0 else (xd - x0) / (x1 - x0)
    yr = 0.5 if y1 == y0 else (y_data - y0) / (y1 - y0)

    # Base horizontal: left/right
    if str(force_side or "").lower() == "left":
        base_dx = -abs(dx);  ha = "right"
    elif str(force_side or "").lower() == "right":
        base_dx = +abs(dx);  ha = "left"
    else:
        # fallback: choose based on where we are
        if xr >= 0.5:
            base_dx = -abs(dx); ha = "right"
        else:
            base_dx = +abs(dx); ha = "left"

    # Base vertical: up/down
    if str(prefer_v or "").lower() == "up":
        base_dy = +abs(dy);  va = "bottom"
    elif str(prefer_v or "").lower() == "down":
        base_dy = -abs(dy);  va = "top"
    else:
        # fallback by position
        if yr >= 0.5:
            base_dy = -abs(dy); va = "top"
        else:
            base_dy = +abs(dy); va = "bottom"

    # Keep inside horizontally (near left/right edges)
    if xr > 0.88 and base_dx > 0:  # too close to right edge but pointing right
        base_dx = -abs(dx); ha = "right"
    if xr < 0.12 and base_dx < 0:  # too close to left edge but pointing left
        base_dx = +abs(dx); ha = "left"

    # Keep inside vertically (near top/bottom edges)
    if yr > 0.90 and base_dy > 0:  # near top
        base_dy = -abs(dy); va = "top"
    if yr < 0.10 and base_dy < 0:  # near bottom
        base_dy = +abs(dy); va = "bottom"

    # Avoid collision with the final value tag (if provided)
    if final_xy is not None and fx_num is not None:
        near_x = abs(xd - fx_num) < 6.0  # ~6 days for date axis
        near_y = abs(y_data - final_xy[1]) / max((y1 - y0), 1e-6) < 0.10
        if near_x and near_y:
            base_dy = -base_dy
            va = "top" if va == "bottom" else "bottom"
            # if still near right edge, flip horizontal too
            if xr > 0.92 and base_dx > 0:
                base_dx = -abs(dx); ha = "right"

    return (base_dx, base_dy), ha, va

# ---decide left/right sides from segment centers in axis units ---
def _segment_center_num(dates, i0: int, i1: int) -> float:
    x0 = _mdates.date2num(dates[i0]) if i0 is not None else None
    x1 = _mdates.date2num(dates[i1]) if i1 is not None else None
    if x0 is None or x1 is None:
        return float("nan")
    return 0.5 * (float(x0) + float(x1))

def _max_drawdown_with_idx(y: np.ndarray) -> Tuple[float, int, int]:
    """
    Returns (max_dd, start_idx, end_idx) where max_dd <= 0 (e.g., -0.57).
    """
    if y.size == 0:
        return 0.0, 0, 0
    peak = y[0]
    peak_i = 0
    max_dd = 0.0
    dd_i0 = dd_i1 = 0
    for i in range(1, len(y)):
        if y[i] > peak:
            peak = y[i]
            peak_i = i
        dd = (y[i] / peak) - 1.0
        if dd < max_dd:
            max_dd = dd
            dd_i0, dd_i1 = peak_i, i
    return float(max_dd), int(dd_i0), int(dd_i1)

def _max_drawup_with_idx(y: np.ndarray) -> Tuple[float, int, int]:
    """
    Returns (max_up, start_idx, end_idx) where max_up >= 0 (e.g., +0.83).
    """
    if y.size == 0:
        return 0.0, 0, 0
    trough = y[0]
    trough_i = 0
    max_up = 0.0
    up_i0 = up_i1 = 0
    for i in range(1, len(y)):
        if y[i] < trough:
            trough = y[i]
            trough_i = i
        up = (y[i] / trough) - 1.0
        if up > max_up:
            max_up = up
            up_i0, up_i1 = trough_i, i
    return float(max_up), int(up_i0), int(up_i1)

# --- compute max drawup that does NOT overlap a locked segment (e.g., Max Drawdown) ---

def _max_drawup_non_overlapping(
    y: np.ndarray,
    lock_range: Optional[Tuple[int, int]] = None,
    *,
    prefer_side: str = "either",  # "either" | "right" | "left"
    tol: float = 1e-12,
) -> Tuple[float, int, int]:

    """
    Return the max drawup (>= 0) on y that does NOT overlap a locked segment.
    Backward-compatible:
      - If lock_range is None, behave exactly like _max_drawup_with_idx(y).
    """
    if lock_range is None:
        return _max_drawup_with_idx(y)

    n = int(len(y))
    if n == 0:
        return 0.0, 0, 0

    i0, i1 = int(lock_range[0]), int(lock_range[1])
    i0 = max(0, min(i0, n - 1))
    i1 = max(0, min(i1, n - 1))
    if i1 < i0:
        i0, i1 = i1, i0

    # strictly non-overlapping sides
    L_lo, L_hi = 0, i0 - 1
    R_lo, R_hi = i1 + 1, n - 1

    def _range_drawup(lo: int, hi: int):
        if hi - lo < 1:
            return 0.0, lo, lo
        v, s, e = _max_drawup_with_idx(y[lo:hi + 1])
        return float(v), lo + int(s), lo + int(e)

    # left / right segments
    upL, sL, eL = (0.0, L_lo, L_lo) if L_hi < L_lo else _range_drawup(L_lo, L_hi)
    upR, sR, eR = (0.0, R_hi, R_hi) if R_hi < R_lo else _range_drawup(R_lo, R_hi)

    if upR > upL:
        return upR, sR, eR
    if upL > 0.0:
        return upL, sL, eL


    ps = (prefer_side or "either").lower()
    if ps == "right":
        if upR > tol: return upR, sR, eR
        if upL > tol: return upL, sL, eL
    elif ps == "left":
        if upL > tol: return upL, sL, eL
        if upR > tol: return upR, sR, eR


    # Fallback: zero-length anchor next to the locked region
    anchor = i0 - 1 if i0 - 1 >= 0 else (i1 + 1 if i1 + 1 < n else i0)
    return 0.0, anchor, anchor

def _days_between(dts: pd.Series, i0: int, i1: int) -> int:
    try:
        d0, d1 = pd.to_datetime(dts.iloc[i0]), pd.to_datetime(dts.iloc[i1])
        return int((d1 - d0).days)
    except Exception:
        return max(0, i1 - i0)

# unified foot-lines helper (place footnote outside the axes, bottom-right) ---
def _add_foot_lines(fig, text: str, *, max_fontsize: int = 6, bottom_pad: float | None = None, ha: str = "right") -> None:
    """
    Place explanatory foot lines at the bottom-right *outside* of the matrix/table axes,
    so they never overlap the main content. Computes bottom margin in figure fraction
    using physical size (inches) and dpi to ensure enough space for multi-line notes.
    """
    lines = 1 if not text else text.count("\n") + 1

    # Physical sizing → required bottom inches for the footnote block.
    fig_h_in = float(fig.get_size_inches()[1])
    dpi = float(fig.dpi or 100.0)
    line_px = float(max_fontsize) * 1.25
    need_in = (lines * line_px) / dpi + 0.08  # +0.08in: 余白の安全域

    need_frac = min(0.45, max(need_in / max(fig_h_in, 1e-6), 0.12))
    pad = max(need_frac, float(bottom_pad) if bottom_pad is not None else 0.0)

    try:
        fig.subplots_adjust(bottom=pad)
    except Exception:
        pass

    fig.text(
        0.995, 0.01,
        text or "",
        ha=ha, va="bottom",
        fontsize=int(max_fontsize), color="#555555"
    )

# =============================================================================
# Font (force a CJK-capable family to avoid missing glyphs)
# =============================================================================
def _set_cjk_font(prefer: Optional[List[str]] = None) -> Optional[str]:
    candidates = prefer or [
        "Yu Gothic UI",
        "Yu Gothic",
        "Meiryo",
        "MS Gothic",  # Windows
        "Hiragino Sans",
        "Hiragino Kaku Gothic ProN",  # macOS
        "Noto Sans CJK JP",
        "Noto Sans JP",
        "IPAexGothic",
        "IPAGothic",  # Common
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for fam in candidates:
        if fam in available:
            rcParams["font.family"] = [fam]
            rcParams["font.sans-serif"] = [
                fam,
                "Noto Sans CJK JP",
                "Yu Gothic UI",
                "Meiryo",
                "MS Gothic",
                "IPAexGothic",
                "IPAGothic",
            ]
            rcParams["axes.unicode_minus"] = False
            return fam
    rcParams["axes.unicode_minus"] = False
    return None


_CHOSEN_CJK = _set_cjk_font()
if not _CHOSEN_CJK:
    print("[viz] CJK font not found; you may see missing-glyph warnings.")


# =============================================================================
# Save (delegate watermark)
# =============================================================================

def save_with_watermarks(filepath: Path, dpi: int = 150, bbox_inches=None) -> None:
    """
    Save the current Matplotlib figure WITHOUT any watermark overlay.
    Watermarking is handled in a separate project; this project intentionally saves clean images.
    """
    filepath.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(filepath, dpi=dpi, bbox_inches=bbox_inches)
    plt.close()


# =============================================================================
# Helpers (robust access)
# =============================================================================
def _infer_ppy(freq: str) -> int:
    f = (freq or "").lower()
    if f in ("", "daily", "day"):
        return 252
    if f in ("monthly", "month"):
        return 12
    if f in ("yearly", "year"):
        return 1
    return 252


def _make_caption(
    start_date,
    end_date,
    freq,
    rebalance,
    missing,
    style,
    amount,
    initial_cap,
    window,
    ppy: Optional[int] = None,
    dca_interval: Optional[str] = None,
) -> str:
    """
    Build a caption line for figures. For DCA, include Total Principal
    computed from (initial_cap + amount * #contrib within window).
    """
    style_norm = (style or "").lower()
    total_principal_str = ""
    if style_norm == "dca" and (ppy is not None) and (window is not None):
        # reuse contrib-mask logic
        inter = dca_interval or "every_period"

        def _step(ppy, inter):
            inter = inter.lower()
            if inter in ("", "every_period"):
                return 1
            if inter == "weekly":
                return max(1, int(round((ppy or 252) / 52)))
            if inter == "monthly":
                return max(1, int(round((ppy or 252) / 12)))
            if inter == "quarterly":
                return max(1, int(round((ppy or 252) / 4)))
            if inter == "yearly":
                return max(1, int(round((ppy or 252) / 1)))
            return 1

        step = _step(ppy, inter)
        mask = np.zeros(int(window), dtype=bool)
        upto = max(0, int(window) - 1)
        mask[0:upto:step] = True  # no contribution on the very last step
        n_contrib = int(mask.sum())
        total_principal = float(initial_cap) + float(amount) * n_contrib
        total_principal_str = f" Total Principal: {total_principal:,.0f}"

    if style_norm == "dca":
        style_str = (
            f"DCA (Init: {initial_cap:,.0f}, {amount:,.0f}/period, "
            f"Window={window}){total_principal_str}"
        )
    else:
        style_str = f"Lump Sum (Window={window})"
    return (
        f"Period: {start_date} to {end_date} Freq: {freq} Rebalance: {rebalance} Missing: {missing}\n"
        f"Investment Style: {style_str}"
    )


def _safe_name(s: str) -> str:
    s2 = re.sub(r"[^A-Za-z0-9.\_ \-]+", "_", str(s))
    s2 = re.sub(r"\s+", "_", s2)
    s2 = re.sub(r"_+", "_", s2).strip("_")
    return s2


def _get_array(d: dict, keys: List[str]) -> np.ndarray:
    for k in keys:
        if k in d and d[k] is not None:
            try:
                return np.asarray(d[k], dtype=float)
            except Exception:
                pass
    return np.asarray([], dtype=float)

def _get_final_value_vector(results: dict, name: str) -> np.ndarray:
    """
    Return per-window Final Value vector for a portfolio. If missing,
    fall back to Return as an order-preserving proxy (1 + return).
    """
    m = results.get(name, {})
    fv = _get_array(m, ["Final_Value"])
    if fv.size > 0:
        return fv.astype(float, copy=False)
    r = _get_array(m, ["Return"])
    return (1.0 + r) if r.size > 0 else np.asarray([], dtype=float)


def _compute_win_rate_matrix(
    portfolio_names: List[str],
    results: dict,
    *,
    tie_policy: str = "exclude"  # "exclude" (default) or "half"
) -> tuple[np.ndarray, np.ndarray]:
    """
    Pairwise win-rate matrix based on per-window Final Value vectors.

    win_rate[i,j] = % of windows where FinalValue_i > FinalValue_j
    Denominator:
      - "exclude": ties are excluded
      - "half":    ties count as 0.5 win each (denominator = #windows)
    """
    n = len(portfolio_names)
    win_rate = np.full((n, n), np.nan, dtype=float)
    denom = np.zeros((n, n), dtype=float)

    vecs = [_get_final_value_vector(results, p) for p in portfolio_names]

    for i in range(n):
        vi = vecs[i]
        if vi.size == 0:
            continue
        for j in range(n):
            vj = vecs[j]
            if vj.size == 0:
                continue
            L = int(min(vi.size, vj.size))
            if L <= 0:
                continue

            a = vi[:L]
            b = vj[:L]
            gt = (a > b)
            eq = (a == b)

            if tie_policy == "half":
                wins = gt.sum(dtype=float) + 0.5 * eq.sum(dtype=float)
                d = float(L)
            else:  # "exclude"
                not_tie = ~eq
                wins = np.logical_and(gt, not_tie).sum(dtype=float)
                d = float(not_tie.sum(dtype=float))

            wr = (wins / d * 100.0) if d > 0 else np.nan
            win_rate[i, j] = wr
            denom[i, j] = d

    return win_rate, denom.astype(int, copy=False)

def _get_scalar(d: dict, keys: List[str], default: float = np.nan) -> float:
    for k in keys:
        if k in d and d[k] is not None:
            v = d[k]
            try:
                arr = np.asarray(v, dtype=float)
                if arr.ndim == 0:
                    return float(arr)
                if arr.size == 0:
                    return float(default)
                return float(np.median(arr))
            except Exception:
                try:
                    return float(v)
                except Exception:
                    continue
    return float(default)


def _has_any_key(results: dict, keys: List[str]) -> bool:
    return any(any(k in results[p] for k in keys) for p in results.keys())


def _contrib_mask_for_window(
    window: int, ppy: int, dca_interval: Optional[str]
) -> np.ndarray:
    inter = (dca_interval or "every_period").lower()
    if inter == "every_period":
        step = 1
    elif inter == "weekly":
        step = max(1, int(round(ppy / 52)))
    elif inter == "monthly":
        step = max(1, int(round(ppy / 12)))
    elif inter == "quarterly":
        step = max(1, int(round(ppy / 4)))
    elif inter == "yearly":
        step = max(1, int(round(ppy / 1)))
    else:
        step = 1
    mask = np.zeros(window, dtype=bool)
    upto = max(0, window - 1)
    mask[0:upto:step] = True
    mask[-1] = False
    return mask


def _year_ticks(L: int, ppy: int):
    """
    Return (ticks, labels) for a step axis with 'Year' markers (Y1, Y2, ...).
    Automatically chooses a coarser step for long windows so labels don't crowd.
    """
    if ppy is None or ppy <= 0:
        ppy = 252

    total_years = max(1, int(round(L / float(ppy))))
    # Choose step size based on total years
    # <=20y: 1y, <=50y: 2y, <=100y: 5y, <=200y: 10y, otherwise: 20y
    if total_years <= 20:
        step_years = 1
    elif total_years <= 50:
        step_years = 2
    elif total_years <= 100:
        step_years = 5
    elif total_years <= 200:
        step_years = 10
    else:
        step_years = 20

    # tick positions are in "period steps"
    y_step = max(1, int(round(ppy * step_years)))
    ticks = np.arange(y_step, L + 1, y_step, dtype=int)

    # Labels like Y1, Y3, Y5 ... according to step_years
    labels = [f"Y{int(round(t / float(ppy)))}" for t in ticks]
    return ticks, labels


def _dca_irr_from_median_fv(
    fv_median: float,
    window: int,
    ppy: int,
    amount: float,
    initial_cap: float,
    dca_interval: Optional[str] = None,
    *,
    max_iter: int = 120,
    tol: float = 1e-10,
) -> float:
    """
    IRR (annualized) solved from median Final Value given the DCA cash-flow schedule.
    Contributions occur at period start; no contribution at last step.
    """
    if not np.isfinite(fv_median) or fv_median <= 0 or window <= 1 or ppy <= 0:
        return np.nan

    inter = (dca_interval or "every_period").lower()
    if inter == "every_period":
        step = 1
    elif inter == "weekly":
        step = max(1, int(round(ppy / 52)))
    elif inter == "monthly":
        step = max(1, int(round(ppy / 12)))
    elif inter == "quarterly":
        step = max(1, int(round(ppy / 4)))
    elif inter == "yearly":
        step = max(1, int(round(ppy / 1)))
    else:
        step = 1

    upto = max(0, window - 1)
    K = np.arange(0, upto, step, dtype=float)
    Y = float(window) / float(ppy)

    I0 = float(initial_cap)
    A = float(amount)

    def f(r: float) -> float:
        base = 1.0 + r
        if base <= 0.0:
            return np.inf
        term0 = I0 * (base**Y)
        termA = 0.0 if A == 0.0 else A * np.sum(base ** (Y - K / float(ppy)))
        return term0 + termA - float(fv_median)

    lo, hi = -0.999, 10.0
    flo, fhi = f(lo), f(hi)
    it_expand = 0
    while np.sign(flo) == np.sign(fhi) and it_expand < 24:
        hi *= 1.5
        fhi = f(hi)
        it_expand += 1
    if np.isnan(flo) or np.isnan(fhi) or np.sign(flo) == np.sign(fhi):
        return np.nan

    for _ in range(max_iter):
        mid = 0.5 * (lo + hi)
        fm = f(mid)
        if not np.isfinite(fm):
            return np.nan
        if abs(fm) < tol or abs(hi - lo) < max(tol, 1e-12):
            return float(mid)
        if np.sign(fm) == np.sign(flo):
            lo, flo = mid, fm
        else:
            hi, fhi = mid, fm
    return float(0.5 * (lo + hi))

# =============================================================================
# Excel export helper for Performance Summary Table + Notes
# =============================================================================
def export_performance_summary_to_excel(
    output_dir: Path,
    filename: str,
    *,
    cols_top: list,
    rows_top: list,
    cols_mid: list,
    rows_mid: list,
    notes_lines: list,
    caption: str,
    meta: dict,
    numeric_payload: Optional[dict] = None,
) -> Path:
    """
    Write a single-sheet Excel workbook that mirrors
    'performance_summary_table_notes.png'. Everything is written into a
    single worksheet named 'Summary' in English.

    Layout:
      A top title -> Top table -> Middle table -> Notes -> Caption -> Parameters

    Number formats are applied per column label where available.
    """
    import pandas as _pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    output_dir.mkdir(parents=True, exist_ok=True)
    xls_path = output_dir / filename

    # Build DataFrames (prefer numeric payload to preserve types)
    if numeric_payload and isinstance(numeric_payload, dict):
        df_top = numeric_payload.get("df_top_num")
        df_mid = numeric_payload.get("df_mid_num")
        top_number_formats = numeric_payload.get("top_number_formats")
        mid_number_formats = numeric_payload.get("mid_number_formats")
    else:
        df_top = None
        df_mid = None
        top_number_formats = None
        mid_number_formats = None

    if df_top is None:
        df_top = _pd.DataFrame(rows_top, columns=cols_top)
    if df_mid is None:
        df_mid = _pd.DataFrame(rows_mid, columns=cols_mid)

    df_params = _pd.DataFrame(
        sorted([(k, v) for k, v in meta.items()], key=lambda x: x[0]),
        columns=["Key", "Value"],
    )

    def _apply_number_formats(ws, header_row_idx: int, start_col_idx: int, number_formats: dict | None):
        if not number_formats:
            return
        # Map header text -> column index within the sheet
        headers = {}
        row = ws[header_row_idx]
        for j, cell in enumerate(row, start=1):
            if j < start_col_idx:
                continue
            val = cell.value
            if val is None:
                continue
            headers[str(val)] = cell.column
        # Apply formats under each matching header
        max_row = ws.max_row
        for label, fmt in number_formats.items():
            col_idx = headers.get(label)
            if not col_idx:
                continue
            for r in range(header_row_idx + 1, max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = fmt

    def _style_header_row(ws, header_row_idx: int, start_col_idx: int, end_col_idx: int):
        header_fill = PatternFill("solid", fgColor="F0F0F0")
        for c in range(start_col_idx, end_col_idx + 1):
            cell = ws.cell(row=header_row_idx, column=c)
            cell.font = Font(bold=True)
            cell.fill = header_fill

    def _auto_column_widths(ws):
        # Best-effort auto-width; clamp to [10, 60]
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                v = cell.value
                l = len(str(v)) if v is not None else 0
                if l > max_len:
                    max_len = l
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    with _pd.ExcelWriter(xls_path, engine="openpyxl", mode="w") as writer:
        # Create the single summary sheet
        wb = writer.book
        # Remove default sheet if present
        if wb.worksheets and wb.worksheets[0].title == 'Sheet' and len(wb.worksheets) == 1:
            wb.remove(wb.worksheets[0])
        ws = wb.create_sheet(title="Summary")
        writer.sheets["Summary"] = ws

        current_row = 1
        # Title
        ws.cell(row=current_row, column=1, value="Performance Summary (Top + Middle + Notes + Parameters)")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 2

        # --- Top table ---
        ws.cell(row=current_row, column=1, value="Top — Distribution Summary")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        df_top.to_excel(writer, sheet_name="Summary", index=False, startrow=current_row - 1, startcol=0)
        header_row_top = current_row
        start_col_top = 1
        end_col_top = start_col_top + len(df_top.columns) - 1
        current_row += len(df_top) + 2

        # --- Middle table ---
        ws.cell(row=current_row, column=1, value="Middle — Path/CAGR/Risk/Sharpe/Drawups (medians across rolling windows)")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        df_mid.to_excel(writer, sheet_name="Summary", index=False, startrow=current_row - 1, startcol=0)
        header_row_mid = current_row
        start_col_mid = 1
        end_col_mid = start_col_mid + len(df_mid.columns) - 1
        current_row += len(df_mid) + 2

        # --- Notes ---
        ws.cell(row=current_row, column=1, value="Notes:")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        for line in notes_lines:
            ws.cell(row=current_row, column=1, value=line)
            current_row += 1
        current_row += 1

        # --- Caption ---
        ws.cell(row=current_row, column=1, value="Caption:")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        # Place caption in merged cells to span a few columns
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=max(end_col_top, end_col_mid))
        ws.cell(row=current_row, column=2, value=caption).alignment = Alignment(wrap_text=True, vertical='top')
        current_row += 2

        # --- Parameters ---
        ws.cell(row=current_row, column=1, value="Parameters:")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        df_params.to_excel(writer, sheet_name="Summary", index=False, startrow=current_row - 1, startcol=0)
        header_row_params = current_row
        end_col_params = 2
        current_row += len(df_params) + 1

        # Freeze pane just below the big title
        ws.freeze_panes = 'A3'

        # Style headers and apply number formats per-block
        _style_header_row(ws, header_row_top, start_col_top, end_col_top)
        _style_header_row(ws, header_row_mid, start_col_mid, end_col_mid)
        _style_header_row(ws, header_row_params, 1, end_col_params)

        _apply_number_formats(ws, header_row_top, start_col_top, top_number_formats)
        _apply_number_formats(ws, header_row_mid, start_col_mid, mid_number_formats)

        # Auto column widths after all content is written
        _auto_column_widths(ws)

    return xls_path
# =============================================================================
# Console tables (English labels)
# =============================================================================
def print_summary_table(
    console: Console,
    portfolio_names: List[str],
    results: dict,
    val: int,
    unit: str,
    style: str,
    amount: float,
    initial_cap: float,
    window: int,
    ppy: Optional[int] = None,
    dca_interval: Optional[str] = None,
    **kwargs,
) -> None:
    if ppy is None:
        ppy = 252
    dca_interval = dca_interval or "every_period"
    style_norm = (style or "").lower()
    is_dca = style_norm == "dca"
    use_dca_value_table = is_dca

    def _principal_from_window() -> float:
        if not is_dca:
            return float(initial_cap if initial_cap and initial_cap > 0 else 1.0)
        mask = _contrib_mask_for_window(window, ppy, dca_interval)
        n_contrib = int(mask.sum())
        return float(initial_cap) + float(amount) * n_contrib

    principal_window = _principal_from_window()

    if use_dca_value_table:
        title = (
            f"Rolling Performance Summary ({val} {unit}) "
            f"[DCA Init: {initial_cap:,.0f} Total Principal: {principal_window:,.0f}]"
        )
        summary_table = Table(title=title)
        cols = [
            "Portfolio",
            "Final(Mean)",
            "Final(Min)",
            "Final(25%)",
            "Final(Med)",
            "Final(75%)",
            "Final(Max)",
            "Path Min(Med)",
            "Path Max(Med)",
            "Tot CAGR(Med)",
            "IRR(Med)",
            "Max Drawup(Med)",
            "Max DD(Med)",
        ]
        for c in cols:
            summary_table.add_column(
                c, justify=("left" if c == "Portfolio" else "right"), no_wrap=True
            )

        for p in portfolio_names:
            m = results.get(p, {})
            fv = _get_array(m, ["Final_Value"])
            fv_med = float(np.median(fv)) if fv.size else np.nan
            path_min = _get_scalar(m, ["Path_Min", "PathMin"])
            path_max = _get_scalar(m, ["Path_Max", "PathMax"])
            max_up = _get_scalar(m, ["Max_Drawup", "MaxDrawup"])
            max_dd = _get_scalar(m, ["Max_DD", "MaxDD"])
            tot_cagr_simple_med = _get_scalar(m, ["CAGR_Simple"], default=np.nan)
            if (
                (not np.isfinite(tot_cagr_simple_med))
                and np.isfinite(fv_med)
                and principal_window > 0
                and window > 0
                and ppy > 0
            ):
                years = float(window) / float(ppy)
                base = max(fv_med / principal_window, 1e-12)
                tot_cagr_simple_med = np.power(base, 1.0 / years) - 1.0

            irr_med = (
                _dca_irr_from_median_fv(
                    fv_median=fv_med,
                    window=window,
                    ppy=ppy,
                    amount=float(amount or 0.0),
                    initial_cap=float(initial_cap or 0.0),
                    dca_interval=(dca_interval or "every_period"),
                )
                if np.isfinite(fv_med)
                else np.nan
            )

            row = [
                p,
                (f"{np.mean(fv):,.0f}" if fv.size else "—"),
                (f"{np.min(fv):,.0f}" if fv.size else "—"),
                (f"{np.percentile(fv, 25):,.0f}" if fv.size else "—"),
                (f"{fv_med:,.0f}" if np.isfinite(fv_med) else "—"),
                (f"{np.percentile(fv, 75):,.0f}" if fv.size else "—"),
                (f"{np.max(fv):,.0f}" if fv.size else "—"),
                (f"{path_min:,.0f}" if np.isfinite(path_min) else "—"),
                (f"{path_max:,.0f}" if np.isfinite(path_max) else "—"),
                (
                    f"{tot_cagr_simple_med:.2%}"
                    if np.isfinite(tot_cagr_simple_med)
                    else "—"
                ),
                (f"{irr_med:.2%}" if np.isfinite(irr_med) else "—"),
                f"{max_up:.2%}" if np.isfinite(max_up) else "—",
                f"{max_dd:.2%}" if np.isfinite(max_dd) else "—",
            ]
            summary_table.add_row(*row)
        console.print(summary_table)
    else:
        title = f"Rolling Performance Summary ({val} {unit}) [Lump Sum metrics (Return%) used]"
        summary_table = Table(title=title)
        cols = [
            "Portfolio",
            "Mean",
            "Min",
            "25%",
            "Median",
            "75%",
            "Max",
            "Path Min(Med)",
            "Path Max(Med)",
            "CAGR(Med)",
            "Risk(Med)",
            "Sharpe(Med)",
            "Max Drawup(Med)",
            "Max DD(Med)",
        ]
        for c in cols:
            summary_table.add_column(
                c, justify=("left" if c == "Portfolio" else "right"), no_wrap=True
            )
        for p in portfolio_names:
            m = results.get(p, {})
            r = _get_array(m, ["Return"])
            path_min = _get_scalar(m, ["Path_Min", "PathMin"])
            path_max = _get_scalar(m, ["Path_Max", "PathMax"])
            cagr_med = _get_scalar(m, ["CAGR"])
            risk_med = _get_scalar(m, ["Risk"])
            shrp_med = _get_scalar(m, ["Sharpe"])
            max_up = _get_scalar(m, ["Max_Drawup", "MaxDrawup"])
            max_dd = _get_scalar(m, ["Max_DD", "MaxDD"])
            summary_table.add_row(
                p,
                (f"{np.mean(r):.2%}" if r.size else "—"),
                (f"{np.min(r):.2%}" if r.size else "—"),
                (f"{np.percentile(r, 25):.2%}" if r.size else "—"),
                (f"{np.median(r):.2%}" if r.size else "—"),
                (f"{np.percentile(r, 75):.2%}" if r.size else "—"),
                (f"{np.max(r):.2%}" if r.size else "—"),
                f"{path_min:.2f}" if np.isfinite(path_min) else "—",
                f"{path_max:.2f}" if np.isfinite(path_max) else "—",
                f"{cagr_med:.2%}" if np.isfinite(cagr_med) else "—",
                f"{risk_med:.2%}" if np.isfinite(risk_med) else "—",
                f"{shrp_med:.2f}" if np.isfinite(shrp_med) else "—",
                f"{max_up:.2%}" if np.isfinite(max_up) else "—",
                f"{max_dd:.2%}" if np.isfinite(max_dd) else "—",
            )
        console.print(summary_table)


# CHANGED: hide "(n=...)" and colorize cells (>50% green, <50% red)
from typing import List
import numpy as np
from rich.console import Console
from rich.table import Table

def print_win_rate_table(
    console: Console,
    portfolio_names: List[str],
    results: dict,
    tie_policy: str = "exclude",
) -> None:
    """
    Print pairwise win-rate matrix based on per-window Final Value vectors.
    Display rule:
      - Value-only percentage (no sample size "n=...").
      - > 50% shown in green, < 50% shown in red, exactly 50% shown in default color.
    """
    win_rate, denom = _compute_win_rate_matrix(
        portfolio_names, results, tie_policy=tie_policy
    )

    tb = Table(title="Win Rate Matrix (Row Final Value > Column Final Value, %)")
    tb.add_column("Portfolio", style="bold")
    for name in portfolio_names:
        tb.add_column(name)

    for i, row_name in enumerate(portfolio_names):
        row_cells = []
        for j, _ in enumerate(portfolio_names):
            if i == j:
                row_cells.append("—")
                continue
            wr = win_rate[i, j]
            if np.isfinite(wr):
                # Colorize: >50% green, <50% red
                style = "green" if wr > 50.0 else ("red" if wr < 50.0 else "")
                text = f"{wr:.1f}%"
                cell = f"[{style}]{text}[/{style}]" if style else text
            else:
                cell = "n/a"
            row_cells.append(cell)
        tb.add_row(row_name, *row_cells)

    console.print(tb)

# =============================================================================
# Charts + CSV exports
# =============================================================================
def save_charts_and_tables(
    output_dir: Path,
    portfolio_names: List[str],
    results: dict,
    val: int,
    unit: str,
    start_date,
    end_date,
    freq,
    rebalance,
    missing,
    style: str,
    amount: float,
    initial_cap: float,
    window: int,
    prices_df: pd.DataFrame,
    dca_interval: Optional[str] = None,
    ppy: Optional[int] = None,
    tax_reports: Optional[dict] = None,
    rolling_mm_df: Optional[pd.DataFrame] = None,
    roll_paths_mean: Optional[dict] = None,
    roll_paths_median: Optional[dict] = None,
    roll_tax_sums: Optional[dict] = None,
    tax_win_start=None,
    tax_win_end=None,
    typical_value_mean: Optional[dict] = None,
    typical_value_median: Optional[dict] = None,
    representative_paths: Optional[dict] = None,
) -> None:
    sns.set_theme(style="whitegrid")
    style_norm = (style or "").lower()
    is_dca = style_norm == "dca"

    caption = _make_caption(
        start_date,
        end_date,
        freq,
        rebalance,
        missing,
        style,
        amount,
        initial_cap,
        window,
        ppy=ppy,
        dca_interval=dca_interval,
    )
    if not ppy:
        ppy = _infer_ppy(freq)
    n_ports = len(portfolio_names)

    # ---------- (1) Return Distribution ----------
    plt.figure(figsize=(10, 6))
    for p in portfolio_names:
        d = _get_array(results.get(p, {}), ["Return"])
        if d.size == 0:
            continue
        sns.histplot(
            d * 100,
            label=p,
            kde=True,
            stat="density",
            common_norm=False,
            alpha=0.5,
            bins=50,
        )
    try:
        overall_min = min(
            float(np.min(_get_array(results.get(p, {}), ["Return"])) * 100)
            for p in portfolio_names
            if _get_array(results.get(p, {}), ["Return"]).size > 0
        )
        if plt.xlim()[0] > overall_min:
            plt.xlim(left=max(overall_min - 10, overall_min * 0.9))
    except ValueError:
        pass
    plt.title(
        f"Rolling Return Distribution ({val} {unit})\n{caption}", fontsize=12, pad=15
    )
    plt.xlabel("Return (%)")
    plt.ylabel("Density")
    plt.legend(loc="upper right")
    plt.tight_layout()
    save_with_watermarks(output_dir / "return_dist_combined.png")

    # ---------- (2) Return Boxplot ----------
    plt.figure(figsize=(10, max(4, n_ports * 1.5)))
    arrs = [
        (_get_array(results.get(p, {}), ["Return"]) * 100)
        for p in portfolio_names
        if _get_array(results.get(p, {}), ["Return"]).size > 0
    ]
    if len(arrs) > 0:
        sns.boxplot(data=arrs, orient="h", palette="Set2", showfliers=False)
        overall_min = min(float(np.min(a)) for a in arrs)
        cur_xlim = plt.xlim()
        plt.xlim(min(overall_min - 15, cur_xlim[0]), cur_xlim[1])
        ys = [
            p
            for p in portfolio_names
            if _get_array(results.get(p, {}), ["Return"]).size > 0
        ]
        for i, p in enumerate(ys):
            mn = float(np.min(_get_array(results.get(p, {}), ["Return"])) * 100)
            plt.scatter(mn, i, color="red", marker="x", s=100, zorder=10)
            plt.text(
                mn,
                i - 0.2,
                f"{mn:.1f}%",
                color="red",
                ha="center",
                va="bottom",
                fontsize=10,
                fontweight="bold",
            )
        plt.yticks(range(len(ys)), ys)
    plt.title(f"Return Boxplot ({val} {unit})\n{caption}", fontsize=12, pad=15)
    plt.xlabel("Return (%)")
    plt.tight_layout()
    save_with_watermarks(output_dir / "return_boxplot.png")

    # ---------- (3) Win Rate Matrix ----------
    # Pairwise Final Value comparison per rolling window (ties excluded by default)
    win_rate, denom = _compute_win_rate_matrix(
        portfolio_names, results, tie_policy="exclude"
    )

    n_ports = len(portfolio_names)
    fig_w = max(8.0, min(1.4 * n_ports + 2.0, 28.0))
    fig_h = max(8.0, min(1.6 * n_ports + 1.5, 24.0))
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))

    # Prepare heatmap values and annotations
    display = np.nan_to_num(win_rate, nan=0.0)  # colormap only
    annot = np.empty_like(display, dtype=object)
    for i in range(n_ports):
        for j in range(n_ports):
            if i == j:
                annot[i, j] = "—"
            else:
                v = win_rate[i, j]
                n_eff = int(denom[i, j])
                annot[i, j] = "n/a" if (not np.isfinite(v) or n_eff <= 0) else f"{v:.1f}%\n(n={n_eff})"

    sns.heatmap(
        display,
        annot=annot,
        fmt="",
        cmap="RdYlGn",
        center=50,
        vmin=0, vmax=100,
        xticklabels=portfolio_names,
        yticklabels=portfolio_names,
        cbar_kws={"label": "Win Rate (%)"},
        ax=ax,
    )

    ax.set_title(
        f"Win Rate Matrix (%) [Row Final Value > Column Final Value]\n{caption}",
        fontsize=12, pad=15
    )

    # --- foot lines (outside, bottom-right; font size <= 6) ---
    foot_lines = [
        "Win Rate Matrix (row vs. column)",
        "• Same rolling horizon only; one comparison per window.",
        "• A 'win' counts when Final_Value_row > Final_Value_col.",
        "• Ties are excluded from the denominator.",
    ]
    if (style or "").lower() == "dca":
        foot_lines += [
            "• Final_Value = portfolio value with start-of-period contributions,",
            f"  (initial={initial_cap:,.0f}, amount={amount:,.0f}, interval={dca_interval or 'every_period'}).",
        ]
    else:
        foot_lines += [
            "• Final_Value = initial_cap × (end_price / start_price).",
            f"  (initial={initial_cap:,.0f}).",
        ]
    _add_foot_lines(fig, "\n".join(foot_lines), max_fontsize=6)

    save_with_watermarks(output_dir / "win_rate_matrix.png", bbox_inches="tight")

    # ---------- (4) Performance Summary Table ----------
    try:
        use_dca_value_table = is_dca

        # ---- Top table (text for PNG) + numeric rows (for Excel) ----
        cols_top = (
            [
                "Portfolio",
                "Final(Mean)",
                "Final(Min)",
                "Final(25%)",
                "Final(Med)",
                "Final(75%)",
                "Final(Max)",
            ]
            if use_dca_value_table
            else ["Portfolio", "Mean", "Min", "25%", "Median", "75%", "Max"]
        )
        rows_top: list[list[str]] = []
        rows_top_num: list[list[Optional[float]]] = []  # numeric values for Excel

        for p in portfolio_names:
            m = results.get(p, {})
            if use_dca_value_table:
                fv = _get_array(m, ["Final_Value"])
                if fv.size == 0:
                    rows_top.append([p] + ["—"] * (len(cols_top) - 1))
                    rows_top_num.append([p, None, None, None, None, None, None])
                else:
                    rows_top.append(
                        [
                            p,
                            f"{np.mean(fv):,.0f}",
                            f"{np.min(fv):,.0f}",
                            f"{np.percentile(fv, 25):,.0f}",
                            f"{np.median(fv):,.0f}",
                            f"{np.percentile(fv, 75):,.0f}",
                            f"{np.max(fv):,.0f}",
                        ]
                    )
                    rows_top_num.append(
                        [
                            p,
                            float(np.mean(fv)),
                            float(np.min(fv)),
                            float(np.percentile(fv, 25)),
                            float(np.median(fv)),
                            float(np.percentile(fv, 75)),
                            float(np.max(fv)),
                        ]
                    )
            else:
                r = _get_array(m, ["Return"])
                if r.size == 0:
                    rows_top.append([p] + ["—"] * (len(cols_top) - 1))
                    rows_top_num.append([p, None, None, None, None, None, None])
                else:
                    rows_top.append(
                        [
                            p,
                            f"{np.mean(r):.2%}",
                            f"{np.min(r):.2%}",
                            f"{np.percentile(r, 25):.2%}",
                            f"{np.median(r):.2%}",
                            f"{np.percentile(r, 75):.2%}",
                            f"{np.max(r):.2%}",
                        ]
                    )
                    rows_top_num.append(
                        [
                            p,
                            float(np.mean(r)),
                            float(np.min(r)),
                            float(np.percentile(r, 25)),
                            float(np.median(r)),
                            float(np.percentile(r, 75)),
                            float(np.max(r)),
                        ]
                    )

        # ---- Middle table (text for PNG) + numeric rows (for Excel) ----
        if use_dca_value_table:
            cols_mid = [
                "Portfolio",
                "Path Min(Med)",
                "Path Max(Med)",
                "Tot CAGR(Simple, Med)",
                "IRR(Med)",
                "Risk(Med)",
                "Sharpe(Med)",
                "Max Drawup(Med)",
                "Max DD(Med)",
            ]
        else:
            cols_mid = [
                "Portfolio",
                "Path Min(Med)",
                "Path Max(Med)",
                "CAGR(Med)",
                "Risk(Med)",
                "Sharpe(Med)",
                "Max Drawup(Med)",
                "Max DD(Med)",
            ]

        rows_mid: list[list[str]] = []
        rows_mid_num: list[list[Optional[float]]] = []

        for p in portfolio_names:
            m = results.get(p, {})
            path_min = _get_scalar(m, ["Path_Min", "PathMin"])
            path_max = _get_scalar(m, ["Path_Max", "PathMax"])
            cagr_med = _get_scalar(m, ["CAGR"])
            risk_med = _get_scalar(m, ["Risk"])
            shrp_med = _get_scalar(m, ["Sharpe"])
            max_up = _get_scalar(m, ["Max_Drawup", "MaxDrawup"])
            max_dd = _get_scalar(m, ["Max_DD", "MaxDD"])

            if not use_dca_value_table:
                rows_mid.append(
                    [
                        p,
                        f"{path_min:.2f}",
                        f"{path_max:.2f}",
                        f"{cagr_med:.2%}",
                        f"{risk_med:.2%}",
                        f"{shrp_med:.2f}",
                        f"{max_up:.2%}",
                        f"{max_dd:.2%}",
                    ]
                )
                rows_mid_num.append(
                    [
                        p,
                        float(path_min) if np.isfinite(path_min) else None,
                        float(path_max) if np.isfinite(path_max) else None,
                        float(cagr_med) if np.isfinite(cagr_med) else None,
                        float(risk_med) if np.isfinite(risk_med) else None,
                        float(shrp_med) if np.isfinite(shrp_med) else None,
                        float(max_up) if np.isfinite(max_up) else None,
                        float(max_dd) if np.isfinite(max_dd) else None,
                    ]
                )
            else:
                fv_med = _get_scalar(m, ["Final_Value"])
                tot_cagr_simple_med = _get_scalar(m, ["CAGR_Simple"], default=np.nan)
                if (not np.isfinite(tot_cagr_simple_med)) and np.isfinite(fv_med):
                    years = float(window) / float(ppy)
                    principal_window = float(initial_cap) + float(amount) * int(
                        _contrib_mask_for_window(
                            window, ppy, dca_interval or "every_period"
                        ).sum()
                    )
                    if principal_window > 0 and years > 0:
                        base = max(fv_med / principal_window, 1e-12)
                        tot_cagr_simple_med = np.power(base, 1.0 / years) - 1.0

                irr_med = (
                    _dca_irr_from_median_fv(
                        fv_median=fv_med,
                        window=window,
                        ppy=ppy,
                        amount=float(amount or 0.0),
                        initial_cap=float(initial_cap or 0.0),
                        dca_interval=(dca_interval or "every_period"),
                    )
                    if np.isfinite(fv_med)
                    else np.nan
                )

                rows_mid.append(
                    [
                        p,
                        (f"{path_min:,.0f}" if np.isfinite(path_min) else "—"),
                        (f"{path_max:,.0f}" if np.isfinite(path_max) else "—"),
                        (
                            f"{tot_cagr_simple_med:.2%}"
                            if np.isfinite(tot_cagr_simple_med)
                            else "—"
                        ),
                        (f"{irr_med:.2%}" if np.isfinite(irr_med) else "—"),
                        f"{risk_med:.2%}" if np.isfinite(risk_med) else "—",
                        f"{shrp_med:.2f}" if np.isfinite(shrp_med) else "—",
                        f"{max_up:.2%}" if np.isfinite(max_up) else "—",
                        f"{max_dd:.2%}" if np.isfinite(max_dd) else "—",
                    ]
                )
                rows_mid_num.append(
                    [
                        p,
                        float(path_min) if np.isfinite(path_min) else None,
                        float(path_max) if np.isfinite(path_max) else None,
                        float(tot_cagr_simple_med)
                        if np.isfinite(tot_cagr_simple_med)
                        else None,
                        float(irr_med) if np.isfinite(irr_med) else None,
                        float(risk_med) if np.isfinite(risk_med) else None,
                        float(shrp_med) if np.isfinite(shrp_med) else None,
                        float(max_up) if np.isfinite(max_up) else None,
                        float(max_dd) if np.isfinite(max_dd) else None,
                    ]
                )

        # ---- build PNG----
        fig_w = max(14.0, min(6.0 + 1.20 * n_ports, 32.0))
        fig_h = 8.5 if n_ports <= 6 else min(8.5 + 0.35 * (n_ports - 6), 14.0)

        # --- Prepare notes text early (to size the bottom panel dynamically) ---
        if use_dca_value_table:
            notes_lines = [
                "Notes (DCA):",
                "• Top — Final Value distribution (currency) across rolling windows.",
                "• Middle — Medians across windows:",
                "  - Path Min/Max (Med): Median of within-window portfolio VALUE min/max "
                "(contributions at period start; none at last step).",
                "  - Tot CAGR (Simple, Med): Principal-based annualized rate with total contributed principal (I0 + A × #contrib).",
                "  - IRR (Med): Money-weighted return solved from the median Final Value and the DCA cash-flow schedule.",
                "  - Risk/Sharpe: Computed on price-path basis; drawup/down from normalized paths.",
                "  - Max DD(Drawdown): Max cumulative fall from rolling normalized path.",
                "  - Max Drawup:Computed on segments that do NOT overlap the Max Drawdown interval "
                "within the SAME rolling window. The Max Drawdown interval is locked FIRST (priority), \n"
                "then Max Drawup is searched only on the LEFT or RIGHT side of that interval; "
                "ties prefer the RIGHT side.\n"
                "This policy ensures Max Drawup starts/ends either BEFORE or AFTER the Max Drawdown."
                "Assumptions: data freq sets ppy; DCA schedule follows '--dca-interval'; last period has no new contribution.",
            ]
        else:
            notes_lines = [
                "Notes (Lump Sum):",
                "• Top — Return distribution (%, rolling windows) across portfolios.",
                "• Middle — Medians across windows:",
                "  - Path Min/Max (Med): Median of within-window normalized path extrema (start=1).",
                "  - CAGR/Risk/Sharpe: Standard annualization with ppy set by '--freq'.",
                "  - Max DD(Drawdown): Max cumulative fall from rolling normalized path.",
                "  - Max Drawup:Computed on segments that do NOT overlap the Max Drawdown interval "
                "within the SAME rolling window. The Max Drawdown interval is locked FIRST (priority), \n"
                "then Max Drawup is searched only on the LEFT or RIGHT side of that interval; "
                "ties prefer the RIGHT side.\n"
                "This policy ensures Max Drawup starts/ends either BEFORE or AFTER the Max Drawdown."
            ]
        notes_text = "\n".join(notes_lines)
        n_note_lines = notes_text.count("\n") + 1
        # Bottom panel height (relative) grows gently with note lines (clamped)
        h_bot = min(3.5, max(1.2, 0.16 * n_note_lines))

        fig = plt.figure(figsize=(fig_w, fig_h))
        gs = fig.add_gridspec(nrows=3, ncols=1, height_ratios=[3.2, 3.2, h_bot])

        # --- Top table ---
        ax_top = fig.add_subplot(gs[0, 0])
        ax_top.axis("off")
        tab_top = ax_top.table(cellText=rows_top, colLabels=cols_top, loc="center", cellLoc="center")
        tab_top.auto_set_font_size(False)
        tab_top.set_fontsize(10)
        tab_top.scale(1.0, 1.20)
        for (i, j), cell in tab_top.get_celld().items():
            if i == 0:
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#f0f0f0")
            elif j == 0:
                cell.set_text_props(weight="bold")
            try:
                fam = (_CHOSEN_CJK or (rcParams.get("font.family") or [""])[0] or "sans-serif")
                cell.get_text().set_fontfamily(fam)
            except Exception:
                pass
        cap_mode = "DCA (values)" if use_dca_value_table else "Returns"
        ax_top.set_title(
            f"Performance Summary — Top: Mean to Max ({val} {unit}) [{cap_mode}]\n{caption}",
            fontsize=12, fontweight="bold", pad=1,
        )

        # --- Middle table ---
        ax_mid = fig.add_subplot(gs[1, 0])
        ax_mid.axis("off")
        tab_mid = ax_mid.table(cellText=rows_mid, colLabels=cols_mid, loc="center", cellLoc="center")
        tab_mid.auto_set_font_size(False)
        tab_mid.set_fontsize(10)
        tab_mid.scale(1.0, 1.20)
        for (i, j), cell in tab_mid.get_celld().items():
            if i == 0:
                cell.set_text_props(weight="bold")
                cell.set_facecolor("#f0f0f0")
            elif j == 0:
                cell.set_text_props(weight="bold")
            try:
                fam = (_CHOSEN_CJK or (rcParams.get("font.family") or [""])[0] or "sans-serif")
                cell.get_text().set_fontfamily(fam)
            except Exception:
                pass
        ax_mid.set_title(
            "Middle: Path Min – MaxDD + (CAGR/Risk/Sharpe/Drawups) — medians across rolling windows",
            fontsize=12, fontweight="bold", pad=1,
        )

        # --- Bottom notes (left-aligned, bottom-left *inside* the bottom panel) ---
        ax_bot = fig.add_subplot(gs[2, 0])
        ax_bot.axis("off")
        ax_bot.text(
            0.0, 0.02,  # a tiny lift from the bottom edge
            notes_text,
            ha="left", va="bottom",
            fontsize=6, color="#555555",
            transform=ax_bot.transAxes,
            linespacing=1.15,
        )

        # Tight layout is safe because notes are inside ax_bot, not figure margins
        fig.tight_layout()
        save_with_watermarks(output_dir / "performance_summary_table_notes.png", bbox_inches="tight")

        # ===== NEW: Excel export (mirrors the same content) =====

        # Meta for reproducibility
        meta = {
            "freq": str(freq),
            "rebalance": str(rebalance),
            "missing": str(missing),
            "style": str(style),
            "dca_interval": str(dca_interval or "every_period"),
            "amount_per_period": float(amount),
            "initial_cap": float(initial_cap),
            "window_periods": int(window),
            "ppy": int(ppy) if ppy else None,
            "unit": str(unit),
            "val": int(val),
            "start_date": str(start_date),
            "end_date": str(end_date),
        }

        # Numeric DataFrames + number formats
        df_top_num = pd.DataFrame(rows_top_num, columns=cols_top)
        df_mid_num = pd.DataFrame(rows_mid_num, columns=cols_mid)

        if use_dca_value_table:
            top_number_formats = {
                "Final(Mean)": "#,##0",
                "Final(Min)": "#,##0",
                "Final(25%)": "#,##0",
                "Final(Med)": "#,##0",
                "Final(75%)": "#,##0",
                "Final(Max)": "#,##0",
            }
            mid_number_formats = {
                "Path Min(Med)": "#,##0",
                "Path Max(Med)": "#,##0",
                "Tot CAGR(Simple, Med)": "0.00%",
                "IRR(Med)": "0.00%",
                "Risk(Med)": "0.00%",
                "Sharpe(Med)": "0.00",
                "Max Drawup(Med)": "0.00%",
                "Max DD(Med)": "0.00%",
            }
        else:
            top_number_formats = {
                "Mean": "0.00%",
                "Min": "0.00%",
                "25%": "0.00%",
                "Median": "0.00%",
                "75%": "0.00%",
                "Max": "0.00%",
            }
            mid_number_formats = {
                "Path Min(Med)": "0.00",
                "Path Max(Med)": "0.00",
                "CAGR(Med)": "0.00%",
                "Risk(Med)": "0.00%",
                "Sharpe(Med)": "0.00",
                "Max Drawup(Med)": "0.00%",
                "Max DD(Med)": "0.00%",
            }

        numeric_payload = {
            "df_top_num": df_top_num,
            "df_mid_num": df_mid_num,
            "top_number_formats": top_number_formats,
            "mid_number_formats": mid_number_formats,
        }

        try:
            export_performance_summary_to_excel(
                output_dir,
                filename="performance_summary_table_notes.xlsx",
                cols_top=cols_top,
                rows_top=rows_top,
                cols_mid=cols_mid,
                rows_mid=rows_mid,
                notes_lines=notes_lines,
                caption=caption,
                meta=meta,
                numeric_payload=numeric_payload,
            )
        except ImportError as e:
            print(
                "[viz] Excel export skipped: 'openpyxl' is not installed. Install it and re-run.",
                e,
            )
        except Exception as e:
            print("[viz] Excel export error:", e)

    except Exception:
        pass

    # ---------- (5) Typical VALUE paths (Mean-only & Median-only) + CSV----------
    try:
        tv_mean = typical_value_mean or {}
        tv_median = typical_value_median or {}
        L = None
        for p in portfolio_names:
            if p in tv_mean:
                L = len(tv_mean[p])
                break
        if L and L > 0:
            step_arr = np.arange(1, L + 1, dtype=int)
            contrib = np.zeros(L, dtype=float)
            if is_dca and amount and amount > 0:
                mask = _contrib_mask_for_window(L, ppy, dca_interval or "every_period")
                contrib[mask] = float(amount)

            # CSV（従来通り）
            df_mean = pd.DataFrame({"Step": step_arr, "ContributionPerPeriod": contrib})
            for p in portfolio_names:
                if p in tv_mean:
                    df_mean[p] = tv_mean[p]
            df_mean.to_csv(
                output_dir / "typical_value_mean_all_ports.csv",
                index=False,
                encoding="utf-8-sig",
            )

            df_median = pd.DataFrame(
                {"Step": step_arr, "ContributionPerPeriod": contrib}
            )
            for p in portfolio_names:
                if p in tv_median:
                    df_median[p] = tv_median[p]
                elif p in tv_mean:
                    df_median[p] = tv_mean[p]
            df_median.to_csv(
                output_dir / "typical_value_median_all_ports.csv",
                index=False,
                encoding="utf-8-sig",
            )

            # ---------- Mean (linear) + final labels ----------
            plt.figure(figsize=(12, 6))
            ax = plt.gca()
            if is_dca and amount and amount > 0:
                mask = _contrib_mask_for_window(L, ppy, dca_interval or "every_period")
                ax2 = ax.twinx()
                bar_x = np.arange(1, L + 1, dtype=int)[mask]
                # Make contribution bars bolder and in front of the grid/lines just enough.
                ax2.bar(
                    bar_x,
                    np.full(bar_x.shape, float(amount)),
                    width=3.0,  # thicker bars (was 1.8)
                    color="#5f5f5f",  # darker neutral
                    alpha=0.55,  # more opaque (was 0.22)
                    edgecolor="#3d3d3d",  # subtle outline
                    linewidth=0.35,
                    label="Contribution (schedule)",
                    zorder=3,  # lift above most lines; still below annotations
                )

                ax2.set_ylabel("Contribution per period")

            for p in portfolio_names:
                if p in tv_mean:
                    # y is the typical normalized price path Pn (start=1)
                    y = np.asarray(tv_mean[p], dtype=float)

                    # --- DCA only: convert Pn -> VALUE (currency) ---
                    if is_dca:
                        # Contribution schedule (start-of-period, no last-step contrib)
                        mask_bool = _contrib_mask_for_window(L, ppy, dca_interval or "every_period")
                        inv_pn = 1.0 / np.clip(y, 1e-12, None)
                        shares = float(initial_cap or 0.0) + float(amount or 0.0) * np.cumsum(
                            inv_pn * mask_bool.astype(float))
                        y_plot = y * shares  # VALUE in currency units
                    else:
                        y_plot = y  # Lump sum: keep as-is (currency if initial_cap>0 else index-like)

                    ax.plot(step_arr, y_plot, label=f"{p} Mean", linewidth=1.9, zorder=5)

                    # final value label (currency-friendly)
                    ax.annotate(
                        f"{y_plot[-1]:,.0f}",
                        xy=(step_arr[-1], y_plot[-1]),
                        xytext=(5, 0),
                        textcoords="offset points",
                        va="center",
                        fontsize=9,
                        color="#333",
                    )

            # Left axis → portfolio amount; format as K/M/B/T
            ax.set_ylabel("Portfolio amount")
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: format_number_kmg(x)))

            # Right axis (contribution) → also format
            if is_dca and amount and amount > 0:
                ax2.set_ylabel("Contribution per period")
                ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: format_number_kmg(x)))

            # Legend merge unchanged
            if is_dca and amount and amount > 0:
                lines, labels = ax.get_legend_handles_labels()
                lines2, labels2 = ax2.get_legend_handles_labels()
                ax.legend(lines + lines2, labels + labels2, loc="best")
            else:
                ax.legend(loc="best")

            plt.tight_layout()
            save_with_watermarks(output_dir / "typical_value_mean_all_ports.png", dpi=150)

            # ---------- Median (linear) + final labels ----------
            plt.figure(figsize=(12, 6))
            ax = plt.gca()
            if is_dca and amount and amount > 0:
                mask = _contrib_mask_for_window(L, ppy, dca_interval or "every_period")
                ax2 = ax.twinx()
                bar_x = np.arange(1, L + 1, dtype=int)[mask]
                # Make contribution bars bolder and in front of the grid/lines just enough.
                ax2.bar(
                    bar_x,
                    np.full(bar_x.shape, float(amount)),
                    width=3.0,  # thicker bars (was 1.8)
                    color="#5f5f5f",  # darker neutral
                    alpha=0.55,  # more opaque (was 0.22)
                    edgecolor="#3d3d3d",  # subtle outline
                    linewidth=0.35,
                    label="Contribution (schedule)",
                    zorder=3,  # lift above most lines; still below annotations
                )

                ax2.set_ylabel("Contribution per period")

            for p in portfolio_names:
                src = None
                if p in tv_median:
                    src = np.asarray(tv_median[p], dtype=float)
                    series_label = f"{p} Median"
                elif p in tv_mean:
                    src = np.asarray(tv_mean[p], dtype=float)
                    series_label = f"{p} Median(n/a→Mean)"
                else:
                    continue

                # src is typical normalized price path Pn (median or fallback)
                y = src

                if is_dca:
                    mask_bool = _contrib_mask_for_window(L, ppy, dca_interval or "every_period")
                    inv_pn = 1.0 / np.clip(y, 1e-12, None)
                    shares = float(initial_cap or 0.0) + float(amount or 0.0) * np.cumsum(
                        inv_pn * mask_bool.astype(float))
                    y_plot = y * shares
                else:
                    y_plot = y

                ax.plot(step_arr, y_plot, label=series_label, linewidth=1.9, linestyle="-", zorder=5)

                ax.annotate(
                    f"{y_plot[-1]:,.0f}",
                    xy=(step_arr[-1], y_plot[-1]),
                    xytext=(5, 0),
                    textcoords="offset points",
                    va="center",
                    fontsize=9,
                    color="#333",
                )

            ax.set_ylabel("Portfolio amount")
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: format_number_kmg(x)))

            if is_dca and amount and amount > 0:
                ax2.set_ylabel("Contribution per period")
                ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: format_number_kmg(x)))

            if is_dca and amount and amount > 0:
                lines, labels = ax.get_legend_handles_labels()
                lines2, labels2 = ax2.get_legend_handles_labels()
                ax.legend(lines + lines2, labels + labels2, loc="best")
            else:
                ax.legend(loc="best")

            plt.tight_layout()
            save_with_watermarks(output_dir / "typical_value_median_all_ports.png", dpi=150)
    except Exception as e:
        print("[viz] Typical paths error:", e)
        pass

    # ---------- (6) Representative percentile paths → STACKED (linear only) + CSV ----------
    try:
        reps = representative_paths or {}
        BLUE = "#1f77b4"  # Principal
        GREEN = "#2ca02c"  # Profit (>=0)
        RED = "#d62728"  # Profit (<0)

        # どの程度の密度なら月次/週次に落とすか
        def _auto_rule(n: int) -> str:
            if n > 600:
                return "ME"  # monthly-end (use 'ME' per requirement)
            if n > 250:
                return "W"  # weekly-end
            return "D"

        for p in portfolio_names:
            paths = reps.get(p, None)
            if not paths:
                continue
            for item in paths:
                label = item.get("label", "Path")
                dates = pd.to_datetime(item["dates"])
                values = np.asarray(item["values"], dtype=float)
                principal = np.asarray(
                    item.get("principal", np.zeros_like(values)), dtype=float
                )

                # native dataframe
                df = pd.DataFrame(
                    {"Date": dates, "Principal": principal, "TotalValue": values}
                ).set_index("Date")
                df["Profit"] = df["TotalValue"] - df["Principal"]

                # choose plotting rule (月次/週次に間引き)
                rule = _auto_rule(len(df))
                # NOTE: Per request, use ("ME","W"); otherwise keep native
                df_plot = (
                    df.resample(rule).last().dropna() if rule in ("ME", "W") else df
                )

                # x & series at plotting resolution
                dt_index: pd.DatetimeIndex = pd.DatetimeIndex(df_plot.index)
                idx = dt_index.to_pydatetime()
                principal_plot = df_plot["Principal"].to_numpy(dtype=float, copy=False)
                profit = df_plot["Profit"].to_numpy(dtype=float, copy=False)
                total_plot = df_plot["TotalValue"].to_numpy(dtype=float, copy=False)

                # --- bar width: 隣接最小間隔×0.75。月次なら約20〜22日、週次なら約4日。 ---
                if len(idx) >= 2:
                    deltas = (
                        np.diff(df_plot.index.values)
                        .astype("timedelta64[D]")
                        .astype(float)
                    )
                    min_gap = float(np.nanmin(deltas))
                    # 安全域：1日〜(min_gap*0.9)にクランプ、既定は0.75倍
                    bar_width_days = min(
                        max(1.0, 0.75 * min_gap), max(1.0, 0.9 * min_gap)
                    )
                else:
                    bar_width_days = 10.0

                pos = profit >= 0
                neg = profit < 0

                # ---------- STACKED BARS (linear only) ----------
                plt.figure(figsize=(16, 8))
                ax = plt.gca()
                ax.set_facecolor("white")  # 背景は白

                # 先にフラグと bottom を必ず定義
                pos = profit >= 0
                neg = profit < 0
                # 損失（neg）のときは元本（青）を profit の下から積み、総額の高さを維持
                base_for_principal = np.where(neg, profit, 0.0)

                # 1) 元本（青）— 先に描画して土台にする
                ax.bar(
                    idx,
                    principal_plot,
                    bottom=base_for_principal,  # neg のときは負の profit を底に
                    color=BLUE,
                    width=bar_width_days,
                    label="Principal",
                    zorder=1,
                    alpha=0.65,  # 視認性のためやや薄め
                    edgecolor="none",
                    linewidth=0.0,
                )

                # 2) 利益（緑, profit>=0）
                if np.any(pos):
                    ax.bar(
                        idx[pos],
                        profit[pos],
                        bottom=principal_plot[pos],
                        color=GREEN,
                        width=bar_width_days,
                        label="Profit",
                        zorder=3,
                        alpha=0.90,
                        edgecolor="none",
                        linewidth=0.0,
                    )

                # 3) 損失（赤, profit<0）— 最後に描いて前面に出す
                if np.any(neg):
                    ax.bar(
                        idx[neg],
                        profit[neg],  # 負の値 → 自然に 0 より下へ伸びる
                        color=RED,
                        width=bar_width_days,
                        label="Profit (loss)",
                        zorder=4,  # 一番前面
                        alpha=1.00,  # 不透明でくっきり
                        edgecolor="#8b0000",  # お好みで輪郭。不要なら "none"
                        linewidth=0.3,
                    )

                # final value label（最終トータル）
                if len(idx) > 0:
                    fv = float(total_plot[-1])
                    ax.annotate(
                        format_integer_commas(fv),
                        xy=(idx[-1], total_plot[-1]),
                        xytext=(8, 6),
                        textcoords="offset points",
                        fontsize=10,
                        color="#222",
                        bbox=dict(
                            boxstyle="round,pad=0.2", fc="white", ec="#888", alpha=0.85
                        ),
                    )

                # dynamic major year ticks for long spans
                if len(idx) >= 2:
                    span_days = (df_plot.index[-1] - df_plot.index[0]).days
                    total_years = max(1, int(round(span_days / 365.25)))
                    if total_years <= 20:
                        base_years = 1
                    elif total_years <= 50:
                        base_years = 2
                    elif total_years <= 100:
                        base_years = 5
                    elif total_years <= 200:
                        base_years = 10
                    else:
                        base_years = 20
                else:
                    base_years = 1

                ax.set_ylabel('Portfolio Value')
                # Use compact money units (K/M/B/T) on the y-axis
                ax.ticklabel_format(axis='y', style='plain', useOffset=False)
                ax.yaxis.set_major_formatter(FuncFormatter(lambda x, pos: format_number_kmg(x)))
                ax.set_title(f"{p} — {label} VALUE Path (calendar)\n{caption}")
                ax.xaxis.set_major_locator(mdates.YearLocator(base=base_years))
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
                # (optional) minor ticks: show yearly ticks between coarse majors when base_years>1
                try:
                    if base_years > 1:
                        ax.xaxis.set_minor_locator(mdates.YearLocator(base=1))
                        ax.tick_params(axis="x", which="minor", length=3, labelsize=0)
                except Exception:
                    pass

                ax.tick_params(axis="x", labelrotation=0)
                ax.grid(True, linewidth=0.4, alpha=0.35)
                ax.legend(loc="best", frameon=True)

                outname_img = f"{_safe_name(p)}_rep_{label.replace('%', 'P').replace(' ', '')}_stacked.png"

                # Use the figure handle so it won't be closed by helper; we control the lifecycle.
                fig = ax.figure
                fig.tight_layout()
                fig.savefig(output_dir / outname_img, dpi=150)

                # --- Draw Max Drawdown / Max Drawup between bar-top points (compute on NATIVE series) ---

                # 1) Compute on native (daily) VALUE series to avoid losing peaks/troughs
                total_native = df["TotalValue"].to_numpy(dtype=float, copy=False)
                dates_native = pd.Series(pd.DatetimeIndex(df.index).to_pydatetime())

                md, dd_i0_n, dd_i1_n = _max_drawdown_with_idx(total_native)
                mu, up_i0_n, up_i1_n = _max_drawup_non_overlapping(
                    total_native, (dd_i0_n, dd_i1_n), prefer_side="either"  # keep definition; 'right' is also ok
                )

                # 2) Debug log (console)
                # try:
                #     print(
                #         f"[viz]{p}[DU/DD] DD {md:.2%} {dates_native.iloc[dd_i0_n].date()}→{dates_native.iloc[dd_i1_n].date()} | "
                #         f"DU {mu:.2%} {dates_native.iloc[up_i0_n].date()}→{dates_native.iloc[up_i1_n].date()}"
                #     )
                # except Exception:
                #     pass

                # 3) Fallback: if DU is zero-length or <=0 (shouldn't happen; but keep robust)
                if not np.isfinite(mu) or mu <= 0.0 or up_i0_n == up_i1_n:
                    if dd_i1_n < len(total_native) - 1:
                        mu = (total_native[-1] / total_native[dd_i1_n]) - 1.0
                        up_i0_n, up_i1_n = dd_i1_n, len(total_native) - 1

                # Clamp indices for safety
                dd_i0_n = int(np.clip(dd_i0_n, 0, len(total_native) - 1))
                dd_i1_n = int(np.clip(dd_i1_n, 0, len(total_native) - 1))
                up_i0_n = int(np.clip(up_i0_n, 0, len(total_native) - 1))
                up_i1_n = int(np.clip(up_i1_n, 0, len(total_native) - 1))

                # Coordinates at native dates (annotation supports arbitrary dates on date-axis)
                dd_x0, dd_y0 = dates_native.iloc[dd_i0_n], float(total_native[dd_i0_n])
                dd_x1, dd_y1 = dates_native.iloc[dd_i1_n], float(total_native[dd_i1_n])
                up_x0, up_y0 = dates_native.iloc[up_i0_n], float(total_native[up_i0_n])
                up_x1, up_y1 = dates_native.iloc[up_i1_n], float(total_native[up_i1_n])

                # Durations
                dd_days = _days_between(dates_native, dd_i0_n, dd_i1_n)
                up_days = _days_between(dates_native, up_i0_n, up_i1_n)

                # (collision guard) keep using the resampled final tag
                final_xy = (dt_index[-1], float(total_plot[-1]))

                # --- draw red (DD) ---
                ax.annotate(
                    "",
                    xy=(dd_x1, dd_y1), xytext=(dd_x0, dd_y0),
                    textcoords="data",
                    arrowprops=dict(arrowstyle="-|>", color="red", lw=2.0),
                    zorder=6,
                )

                # Decide relative sides (DU vs DD)
                du_c = _segment_center_num(dates_native, up_i0_n, up_i1_n)
                dd_c = _segment_center_num(dates_native, dd_i0_n, dd_i1_n)
                if not np.isfinite(du_c) or not np.isfinite(dd_c):
                    du_side, dd_side = "right", "left"
                else:
                    du_side = "left" if (du_c <= dd_c) else "right"
                    dd_side = "right" if du_side == "left" else "left"
                du_v, dd_v = "up", "down"

                # Anchors (robust fallback if smart anchor fails)
                try:
                    (dd_offset, dd_ha, dd_va) = _smart_label_anchor(
                        ax, dd_x1, dd_y1, final_xy=final_xy, dx=12, dy=12,
                        force_side=dd_side, prefer_v=dd_v,
                    )
                except Exception:
                    (dd_offset, dd_ha, dd_va) = ((-12, -12), "right", "top")

                try:
                    (up_offset, up_ha, up_va) = _smart_label_anchor(
                        ax, up_x1, up_y1, final_xy=final_xy, dx=12, dy=12,
                        force_side=du_side, prefer_v=du_v,
                    )
                except Exception:
                    (up_offset, up_ha, up_va) = ((12, 12), "left", "bottom")

                # Labels
                ax.annotate(
                    f"▼ Max Drawdown: {md:.2%}, {dd_days} days",
                    xy=(dd_x1, dd_y1), xytext=dd_offset,
                    textcoords="offset points",
                    ha=dd_ha, va=dd_va,
                    color="red", fontsize=11, fontweight="bold",
                    bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="red", alpha=0.95),
                    zorder=7,
                )

                # --- draw green (DU) ---
                ax.annotate(
                    "",
                    xy=(up_x1, up_y1), xytext=(up_x0, up_y0),
                    textcoords="data",
                    arrowprops=dict(arrowstyle="-|>", color="green", lw=2.0),
                    zorder=6,
                )
                ax.annotate(
                    f"▲ Max Drawup: {mu:.2%}, {up_days} days",
                    xy=(up_x1, up_y1), xytext=up_offset,
                    textcoords="offset points",
                    ha=up_ha, va=up_va,
                    color="green", fontsize=11, fontweight="bold",
                    bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="green", alpha=0.95),
                    zorder=7,
                )

                # Save decorated variant
                annotated_name = outname_img.replace(".png", "_annotated.png")
                fig.tight_layout()
                fig.savefig(output_dir / annotated_name, dpi=150)
                import matplotlib.pyplot as _plt
                _plt.close(fig)

                # ---------- CSV（従来のネイティブ＋プロット後の2種） ----------
                cmask = item.get("contrib_mask", None)
                contrib_made = (
                    cmask.astype(int)
                    if isinstance(cmask, np.ndarray)
                    else np.zeros(len(values), dtype=int)
                )
                df_native = pd.DataFrame(
                    {
                        "Date": pd.to_datetime(dates),
                        "Principal": principal,
                        "Profit": values - principal,
                        "TotalValue": values,
                        "ContributionMade(0/1)": contrib_made,
                        "Label": label,
                        "Portfolio": p,
                    }
                )
                outname_csv_native = f"rep_{_safe_name(p)}_{label.replace('%', 'P').replace(' ', '')}_value_path.csv"
                df_native.to_csv(
                    output_dir / outname_csv_native, index=False, encoding="utf-8-sig"
                )

                df_plot_out = df_plot.reset_index()
                df_plot_out.insert(0, "Portfolio", p)
                df_plot_out.insert(1, "Label", label)
                df_plot_out["PlotResampleRule"] = rule
                df_plot_out["BarWidthDays"] = bar_width_days
                outname_csv_plot = f"rep_{_safe_name(p)}_{label.replace('%', 'P').replace(' ', '')}_value_path_plot.csv"
                df_plot_out.rename(columns={"Date": "Date"}).to_csv(
                    output_dir / outname_csv_plot, index=False, encoding="utf-8-sig"
                )
    except Exception as e:
        print("[viz] Representative stacked bars error:", e)
        pass

    # ---------- (7) Taxes: summary + time series (unchanged) ----------
    try:
        if tax_reports and isinstance(tax_reports, dict):
            cols = [
                "Portfolio",
                "Tax Rate",
                "Events",
                "Total Tax",
                "Avg per event",
                "Max single event",
            ]
            data_rows = []
            for p in portfolio_names:
                rep = tax_reports.get(p, {})
                rate = rep.get("tax_rate", None)
                rate_str = (
                    f"{float(rate):.2%}" if isinstance(rate, (int, float)) else "—"
                )
                events = rep.get("tax_events", 0)
                total = rep.get("total_tax", None)
                avg_ev = rep.get("avg_tax_per_event", None)
                max_one = rep.get("max_tax_single_event", None)
                if int(events or 0) <= 0:
                    data_rows.append(
                        [p, rate_str, "0 (no taxable sells)", "—", "—", "—"]
                    )
                else:
                    data_rows.append(
                        [
                            p,
                            rate_str,
                            f"{int(events)}",
                            (f"{float(total):.6f}" if total is not None else "—"),
                            (f"{float(avg_ev):.6f}" if avg_ev is not None else "—"),
                            (f"{float(max_one):.6f}" if max_one is not None else "—"),
                        ]
                    )

            fig, ax = plt.subplots(figsize=(14, 1.2 + 0.5 * max(1, len(data_rows))))
            ax.axis("off")
            tab = ax.table(
                cellText=data_rows, colLabels=cols, loc="center", cellLoc="center"
            )
            tab.auto_set_font_size(False)
            tab.set_fontsize(10)
            tab.scale(1.0, 1.35)
            for (i, j), cell in tab.get_celld().items():
                if i == 0:
                    cell.set_text_props(weight="bold")
                    cell.set_facecolor("#f0f0f0")
                elif j == 0:
                    cell.set_text_props(weight="bold")
                try:
                    fam = (
                        _CHOSEN_CJK
                        or (rcParams.get("font.family") or [""])[0]
                        or "sans-serif"
                    )
                    cell.get_text().set_fontfamily(fam)
                except Exception:
                    pass
            fig.suptitle(
                "Tax Summary (at Rebalance Events)", fontsize=12, fontweight="bold"
            )
            fig.text(
                0.5,
                0.03,
                "Rows with no taxable rebalancing activity are displayed as '—'.",
                ha="center",
                va="bottom",
                fontsize=9,
                fontfamily=(
                    _CHOSEN_CJK
                    or (rcParams.get("font.family") or [""])[0]
                    or "sans-serif"
                ),
            )
            fig.tight_layout(rect=(0, 0.05, 1, 0.95))
            save_with_watermarks(
                output_dir / "tax_summary_table.png", bbox_inches="tight"
            )

            n = len(portfolio_names)
            fig2, axes = plt.subplots(
                nrows=n, ncols=1, figsize=(14, max(2.5, 2.2 * n)), sharex=True
            )
            if n == 1:
                axes = [axes]
            f = str(freq or "").lower()
            for ax, p in zip(axes, portfolio_names):
                rep = tax_reports.get(p, {})
                s = rep.get("tax_paid_series", None)
                if s is None:
                    ax.set_title(f"{p}: (no tax series)")
                    ax.set_ylabel("Tax")
                    continue
                s = s.copy()
                s.index = pd.to_datetime(s.index)
                if f in ("", "daily", "day"):
                    s_plot = s.resample("ME").sum()
                    subtitle = "(monthly sum, daily data)"
                else:
                    s_plot = s
                    subtitle = ""
                ax.bar(
                    s_plot.index,
                    s_plot.values,
                    width=20 if f in ("", "daily", "day") else 10,
                    color="#cc6666",
                )
                ax.set_title(f"{p} — Taxes Paid Over Time {subtitle}")
                ax.set_ylabel("Tax")
                ax.grid(True, axis="y", alpha=0.3)
            axes[-1].set_xlabel("Date")
            plt.tight_layout()
            save_with_watermarks(output_dir / "tax_paid_time_series.png")
    except Exception:
        pass

    # ---------- (8) Taxes: rolling-window table (unchanged) ----------
    try:
        if roll_tax_sums and isinstance(roll_tax_sums, dict):
            rows = []
            cols = [
                "Portfolio",
                "Tax Sum (Med)",
                "Tax Sum (25%)",
                "Tax Sum (75%)",
                "Tax Sum (Min)",
                "Tax Sum (Max)",
            ]
            for p in portfolio_names:
                v = _get_array({"v": roll_tax_sums.get(p, np.array([]))}, ["v"])
                ev = 0
                if tax_reports and isinstance(tax_reports, dict):
                    rep = tax_reports.get(p, {})
                    try:
                        ev = int(rep.get("tax_events", 0) or 0)
                    except Exception:
                        ev = 0
                if ev == 0 or v.size == 0 or np.allclose(v, 0.0):
                    rows.append([p, "—", "—", "—", "—", "—"])
                    continue
                rows.append(
                    [
                        p,
                        f"{np.median(v):.6f}",
                        f"{np.percentile(v, 25):.6f}",
                        f"{np.percentile(v, 75):.6f}",
                        f"{np.min(v):.6f}",
                        f"{np.max(v):.6f}",
                    ]
                )
            plt.figure(figsize=(14, 1.2 + 0.5 * max(1, len(rows))))
            ax = plt.gca()
            ax.axis("off")
            tab = ax.table(
                cellText=rows, colLabels=cols, loc="center", cellLoc="center"
            )
            tab.auto_set_font_size(False)
            tab.set_fontsize(10)
            tab.scale(1.0, 1.35)
            for (i, j), cell in tab.get_celld().items():
                if i == 0:
                    cell.set_text_props(weight="bold")
                    cell.set_facecolor("#f0f0f0")
                elif j == 0:
                    cell.set_text_props(weight="bold")
                try:
                    fam = (
                        _CHOSEN_CJK
                        or (rcParams.get("font.family") or [""])[0]
                        or "sans-serif"
                    )
                    cell.get_text().set_fontfamily(fam)
                except Exception:
                    pass
            main_title = f"Rolling Taxes Summary (window={window} periods)"
            ax.set_title(main_title, fontsize=12, fontweight="bold", pad=10)
            ax.text(
                0.5,
                0.02,
                "Note: '—' indicates no taxable rebalancing sells in the period (or across rolling windows).",
                ha="center",
                va="bottom",
                fontsize=9,
                transform=ax.transAxes,
                fontfamily=(
                    _CHOSEN_CJK
                    or (rcParams.get("font.family") or [""])[0]
                    or "sans-serif"
                ),
            )
            plt.tight_layout()
            save_with_watermarks(
                output_dir / "rolling_tax_summary_table.png", bbox_inches="tight"
            )
    except Exception:
        pass
